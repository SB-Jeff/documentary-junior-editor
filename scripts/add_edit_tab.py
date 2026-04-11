#!/usr/bin/env python3
"""
Add a Paper Cut tab to an existing quotes workbook.

This script takes the selected and sequenced quotes and adds them as a new
"Paper Cut" tab in the Excel workbook. This tab becomes the input for
generate_fcpxml.py.

PAPER CUT TAB FORMAT:
  Seq #, Quote #, Speaker, Section, Quote, Start TC, End TC, Notes

The edits list should contain quotes in the final narrative sequence order.
Each entry is:
  [seq_num, quote_num, speaker_name, section_name, quote_text, start_tc, end_tc, notes]

NOTES:
  - quote_num can be a string like "82a" for split quotes
  - section_name groups quotes into narrative sections (Opening, Challenge, Solution, Impact)
  - The sequence order in this list IS the order they'll appear in the edit
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def add_paper_cut_tab(excel_path, tab_name, edits):
    """
    Add a Paper Cut tab to an existing workbook.

    Args:
        excel_path: path to the existing .xlsx file
        tab_name: name for the new tab (e.g., "Paper Cut - Rob & Dave")
        edits: list of [seq_num, quote_num, speaker, section, quote, start_tc, end_tc, notes]
    """
    wb = load_workbook(excel_path)
    ws = wb.create_sheet(tab_name)

    headers = ["Seq #", "Quote #", "Speaker", "Section", "Quote", "Start TC", "End TC", "Notes"]
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(name="Arial", bold=True, size=10)
    cell_font = Font(name="Arial", size=10)
    section_fill = PatternFill("solid", fgColor="E2EFDA")
    section_font = Font(name="Arial", bold=True, size=10)
    thin_border = Border(bottom=Side(style='thin', color='D9D9D9'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A2'

    current_section = ""
    for row_data in edits:
        row_num = row_data[0] + 1
        section = row_data[3]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal='center')
            if col == 4:
                cell.alignment = Alignment(horizontal='center')
                if section != current_section:
                    cell.font = section_font
                    cell.fill = section_fill
            if col == 5:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col == 8:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        current_section = section

    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 70
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 35

    wb.save(excel_path)
    print(f"Added '{tab_name}' tab with {len(edits)} quotes")


# ============================================================================
# PROJECT-SPECIFIC EDIT LIST (replace for each new project)
# ============================================================================

if __name__ == '__main__':
    # Example usage:
    # edits = [
    #     [1, 41, "Rob Manion", "Opening", "Quote text here...", "00:01:26:00", "00:01:55:00", ""],
    #     [2, 43, "Rob Manion", "Opening", "Quote text here...", "00:02:56:00", "00:03:12:00", ""],
    # ]
    # add_paper_cut_tab("quotes.xlsx", "Paper Cut - Main Edit", edits)
    print("Define edits list and uncomment the function call above.")
