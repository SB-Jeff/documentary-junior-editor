#!/usr/bin/env python3
"""
Generate the All Quotes spreadsheet tab.

This is a TEMPLATE script. The quotes list below is from the Epicor/Manion's project
and should be replaced with quotes from the current project.

QUOTE FORMAT:
Each quote is a list with these fields:
  [quote_number, speaker_name, quote_text, start_timecode, end_timecode, part_number, rationale]

  - quote_number: Sequential integer, grouped by speaker
  - speaker_name: Full name as it appears in the FCPXML
  - quote_text: VERBATIM text from transcript — never paraphrase or edit
  - start_timecode: Approximate start time (HH:MM:SS:FF)
  - end_timecode: Approximate end time (HH:MM:SS:FF)
  - part_number: Which part of the narrative structure (1=Challenge, 2=Solution, 3=Impact)
  - rationale: Brief editorial note on why this quote matters

IMPORTANT RULES:
  - Quotes must be VERBATIM from the transcript
  - You can trim quotes (cut beginning or end) but never change words
  - You can split a quote into parts (e.g., #82a and #82b) for different sections
  - You can reorder quotes freely — chronological interview order doesn't matter
  - New quotes can be pulled from raw transcripts and assigned the next available number
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_quotes_workbook(quotes, output_path):
    """
    Create an Excel workbook with all tagged quotes.

    Args:
        quotes: list of [quote_num, speaker, quote_text, start_tc, end_tc, part, rationale]
        output_path: where to save the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "All Quotes"

    # Headers
    headers = ["#", "Speaker", "Quote", "Start TC", "End TC", "Part", "Rationale", "Selected", "Sequence #"]
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(name="Arial", bold=True, size=10)
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Write data
    for row_data in quotes:
        row_num = row_data[0] + 1  # +1 for header
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = cell_font
            cell.border = thin_border
            if col == 6:  # Part column
                cell.alignment = Alignment(horizontal='center')
            if col == 3:  # Quote column - wrap text
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col == 7:  # Rationale - wrap text
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    ws.column_dimensions['A'].width = 5    # #
    ws.column_dimensions['B'].width = 16   # Speaker
    ws.column_dimensions['C'].width = 70   # Quote
    ws.column_dimensions['D'].width = 14   # Start TC
    ws.column_dimensions['E'].width = 14   # End TC
    ws.column_dimensions['F'].width = 7    # Part
    ws.column_dimensions['G'].width = 55   # Rationale
    ws.column_dimensions['H'].width = 10   # Selected
    ws.column_dimensions['I'].width = 12   # Sequence #

    # Save
    wb.save(output_path)
    print(f"Saved {len(quotes)} quotes to {output_path}")
    return wb


# ============================================================================
# PROJECT-SPECIFIC QUOTES (replace for each new project)
# ============================================================================

# Example from Epicor/Manion's project — replace with your project's quotes
EXAMPLE_QUOTES = [
    # [quote_number, "Speaker Name", "Verbatim quote text", "HH:MM:SS:FF", "HH:MM:SS:FF", part_number, "Rationale"],
]


if __name__ == '__main__':
    if EXAMPLE_QUOTES:
        create_quotes_workbook(EXAMPLE_QUOTES, "quotes.xlsx")
    else:
        print("No quotes defined. Add quotes to EXAMPLE_QUOTES list and re-run.")
